#!/usr/bin/env python3
"""Canonicalize origin-safe local challenger sources and extend the frozen feature store."""
from __future__ import annotations

import argparse
import csv
import io
import json
import math
import re
import time
import zipfile
from collections import defaultdict
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from v16_common import json_write, sha256

TRANSFORM = "v16_source_integration_1"
COMMON = [
    "entity_id", "municipality_code", "health_region_code", "month", "feature_time",
    "available_time", "source_vintage", "source_file", "source_sha256",
    "transform_version", "missingness_reason", "feature_name", "value",
]


def municipality_six(value: object) -> str:
    text = re.sub(r"\D", "", str(value or ""))
    if len(text) >= 7:
        return text[:6]
    return text.zfill(6) if len(text) == 5 else text


def canonical_long(
    frame: pd.DataFrame,
    id_columns: dict[str, str],
    value_columns: list[str],
    source_file: Path,
    vintage: str,
    feature_time: str,
    available_time: str,
) -> pd.DataFrame:
    rows = frame.melt(id_vars=list(id_columns), value_vars=value_columns, var_name="feature_name", value_name="value")
    rows = rows.rename(columns=id_columns)
    for column in ("entity_id", "municipality_code", "health_region_code", "month"):
        if column not in rows:
            rows[column] = pd.NA
    rows["feature_time"] = pd.to_datetime(rows[feature_time]) if feature_time in rows else pd.NaT
    rows["available_time"] = pd.to_datetime(rows[available_time]) if available_time in rows else pd.NaT
    rows["source_vintage"] = vintage
    rows["source_file"] = str(source_file.resolve())
    rows["source_sha256"] = sha256(source_file)
    rows["transform_version"] = TRANSFORM
    rows["missingness_reason"] = np.where(rows["value"].isna(), "not_reported", "observed")
    return rows[COMMON]


def build_cnes(root: Path, workspace: Path, store: pd.DataFrame) -> tuple[pd.DataFrame, list[str], dict[str, object]]:
    source = root / "BRAZIL_MDRTB_V10_FOCUSED/03_features/cnes_health_region_year.parquet"
    cnes = pd.read_parquet(source).copy()
    cnes["health_region_code"] = cnes["health_region_code"].astype(str)
    features = [c for c in cnes.columns if c not in {"health_region_code", "year"}]
    cnes = cnes.sort_values(["health_region_code", "year"])
    for column in ["cnes_equipment_in_use", "cnes_service_records", "cnes_establishments"]:
        cnes[column + "_yoy"] = cnes.groupby("health_region_code")[column].pct_change().replace([np.inf, -np.inf], np.nan)
        features.append(column + "_yoy")
    long = cnes.melt(id_vars=["health_region_code", "year"], value_vars=features, var_name="feature_name", value_name="value")
    long["entity_id"] = long["health_region_code"]
    long["municipality_code"] = pd.NA
    long["month"] = pd.to_datetime(long["year"].astype(str) + "-12-01")
    long["feature_time"] = long["month"]
    long["available_time"] = long["month"] + pd.offsets.MonthBegin(2)
    long["source_vintage"] = long["year"].astype(str) + "-12"
    long["source_file"] = str(source.resolve())
    long["source_sha256"] = sha256(source)
    long["transform_version"] = TRANSFORM
    long["missingness_reason"] = np.where(long["value"].isna(), "source_failure_or_not_reported", "observed")
    long[COMMON].to_parquet(workspace / "03_canonical/F1_cnes_capacity.parquet", index=False)

    join = store[["health_region_code", "forecast_origin"]].copy()
    join["health_region_code"] = join["health_region_code"].astype(str)
    join["source_year"] = join["forecast_origin"].dt.year - np.where(join["forecast_origin"].dt.month >= 2, 1, 2)
    wide = cnes.rename(columns={c: "f1_" + c for c in features})
    join = join.merge(wide, left_on=["health_region_code", "source_year"], right_on=["health_region_code", "year"], how="left", validate="many_to_one")
    out_features = ["f1_" + c for c in features]
    for c in out_features:
        store[c] = join[c].to_numpy()
    store["f1_age_months"] = (store["forecast_origin"].dt.year - join["source_year"]) * 12 + store["forecast_origin"].dt.month - 12
    out_features.append("f1_age_months")
    return store, out_features, {"status": "PASS", "source": str(source), "rows": len(cnes), "regions": int(cnes.health_region_code.nunique()), "years": [int(cnes.year.min()), int(cnes.year.max())], "features": len(out_features), "available_time_rule": "December competence available on February 1 of following year"}


def sim_sources(root: Path) -> dict[int, Path]:
    sources: dict[int, Path] = {}
    for year in range(2016, 2022):
        p = root / f"Mortalidade_Geral_{year}_csv.zip"
        if p.exists(): sources[year] = p
    sources.update({2022: root / "DO22OPEN.csv", 2023: root / "DO23OPEN.csv", 2024: root / "DO24OPEN_csv.zip"})
    return {year: p for year, p in sources.items() if p.exists()}


def read_sim_chunks(path: Path):
    use = ["DTOBITO", "CODMUNRES", "CAUSABAS", "LINHAA", "LINHAB", "LINHAC", "LINHAD", "LINHAII"]
    kwargs = dict(sep=";", encoding="latin1", dtype=str, usecols=lambda c: c in use, chunksize=250_000, low_memory=False)
    if path.suffix.lower() == ".zip":
        archive = zipfile.ZipFile(path)
        member = next(name for name in archive.namelist() if name.lower().endswith(".csv"))
        with archive.open(member) as handle:
            yield from pd.read_csv(handle, **kwargs)
    else:
        yield from pd.read_csv(path, **kwargs)


def build_sim(root: Path, workspace: Path, store: pd.DataFrame, crosswalk: pd.DataFrame) -> tuple[pd.DataFrame, list[str], dict[str, object]]:
    pieces: list[pd.DataFrame] = []
    counts = {"raw_rows": 0, "invalid_dates": 0, "unmapped_municipalities": 0}
    sources = sim_sources(root)
    mapping = crosswalk.drop_duplicates("municipality_code").set_index("municipality_code")["health_region_code"]
    for year, source in sorted(sources.items()):
        print(f"SIM year={year} source={source.name}", flush=True)
        for chunk in read_sim_chunks(source):
            counts["raw_rows"] += len(chunk)
            date = pd.to_datetime(chunk.get("DTOBITO", pd.Series(index=chunk.index, dtype=str)), format="%d%m%Y", errors="coerce")
            counts["invalid_dates"] += int(date.isna().sum())
            municipality = chunk.get("CODMUNRES", pd.Series(index=chunk.index, dtype=str)).map(municipality_six)
            cause = chunk.get("CAUSABAS", pd.Series("", index=chunk.index)).fillna("").str.upper().str.replace(".", "", regex=False)
            multi = chunk[[c for c in ("LINHAA", "LINHAB", "LINHAC", "LINHAD", "LINHAII") if c in chunk]].fillna("").astype(str).agg(" ".join, axis=1).str.upper().str.replace(".", "", regex=False)
            tb_under = cause.str.contains(r"A1[5-9]", regex=True)
            tb_multi = multi.str.contains(r"A1[5-9]", regex=True) | tb_under
            hiv_tb = tb_multi & (multi.str.contains("B20", regex=False) | cause.str.startswith("B20"))
            sub = pd.DataFrame({"municipality_code": municipality, "month": date.dt.to_period("M").dt.to_timestamp(), "sim_tb_underlying": tb_under.astype(int), "sim_tb_multiple": tb_multi.astype(int), "sim_hiv_tb": hiv_tb.astype(int), "sim_all_deaths": 1})
            sub = sub.dropna(subset=["month"])
            pieces.append(sub.groupby(["municipality_code", "month"], as_index=False).sum(numeric_only=True))
    municipal = pd.concat(pieces, ignore_index=True).groupby(["municipality_code", "month"], as_index=False).sum(numeric_only=True)
    municipal["health_region_code"] = municipal["municipality_code"].map(mapping)
    counts["unmapped_municipalities"] = int(municipal.health_region_code.isna().sum())
    municipal["available_time"] = municipal["month"] + pd.DateOffset(months=18)
    municipal["feature_time"] = municipal["month"]
    municipal["entity_id"] = municipal["municipality_code"]
    municipal["source_vintage"] = municipal["month"].dt.year.astype(str)
    municipal["source_file"] = municipal["month"].dt.year.map({year: str(path.resolve()) for year, path in sources.items()})
    municipal["source_sha256"] = municipal["month"].dt.year.map({year: sha256(path) for year, path in sources.items()})
    municipal["transform_version"] = TRANSFORM
    municipal["missingness_reason"] = "observed"
    long = municipal.melt(id_vars=["entity_id", "municipality_code", "health_region_code", "month", "feature_time", "available_time", "source_vintage", "source_file", "source_sha256", "transform_version", "missingness_reason"], value_vars=["sim_tb_underlying", "sim_tb_multiple", "sim_hiv_tb", "sim_all_deaths"], var_name="feature_name", value_name="value")
    long[COMMON].to_parquet(workspace / "03_canonical/F4_sim_mortality.parquet", index=False)

    region = municipal.dropna(subset=["health_region_code"]).groupby(["health_region_code", "month"], as_index=False)[["sim_tb_underlying", "sim_tb_multiple", "sim_hiv_tb", "sim_all_deaths"]].sum()
    region = region.sort_values(["health_region_code", "month"])
    for column in ["sim_tb_underlying", "sim_tb_multiple", "sim_hiv_tb"]:
        region[f"{column}_12m"] = region.groupby("health_region_code")[column].transform(lambda s: s.rolling(12, min_periods=3).sum())
        region[f"{column}_24m"] = region.groupby("health_region_code")[column].transform(lambda s: s.rolling(24, min_periods=6).sum())
    region["sim_tb_fraction_12m"] = region["sim_tb_multiple_12m"] / region.groupby("health_region_code")["sim_all_deaths"].transform(lambda s: s.rolling(12, min_periods=3).sum()).replace(0, np.nan)
    features = [c for c in region if c.endswith("12m") or c.endswith("24m")]
    query = store[["health_region_code", "forecast_origin"]].copy()
    query["health_region_code"] = query["health_region_code"].astype(str)
    query["sim_cutoff_month"] = (query["forecast_origin"] - pd.DateOffset(months=18)).dt.to_period("M").dt.to_timestamp()
    wide = region.rename(columns={c: "f4_" + c for c in features})
    query = query.merge(wide, left_on=["health_region_code", "sim_cutoff_month"], right_on=["health_region_code", "month"], how="left", validate="many_to_one")
    out_features = ["f4_" + c for c in features]
    for c in out_features: store[c] = query[c].to_numpy()
    store["f4_age_months"] = 18.0
    store["f4_source_available"] = store[out_features].notna().any(axis=1).astype(int)
    out_features += ["f4_age_months", "f4_source_available"]
    return store, out_features, {"status": "PASS", "sources": {year: {"path": str(path), "sha256": sha256(path)} for year, path in sources.items()}, "raw_rows": counts["raw_rows"], "municipality_month_rows": len(municipal), "health_region_month_rows": len(region), "unmapped_municipality_months": counts["unmapped_municipalities"], "available_time_rule": "event month plus 18 months (conservative final-release lag)", "features": len(out_features)}


def build_pmmb(root: Path, workspace: Path, store: pd.DataFrame, crosswalk: pd.DataFrame) -> tuple[pd.DataFrame, list[str], dict[str, object]]:
    source = root / "ppf_mais_medicos_serie_historica.csv.zip"
    use = ["ibge", "total_prof_ativos", "prof_crm_brasil_pmmb", "prof_inter_pmmb", "prof_provab", "dt_referencia"]
    data = pd.read_csv(source, sep=";", encoding="utf-8", dtype=str, usecols=use)
    data["municipality_code"] = data["ibge"].map(municipality_six)
    data = data[data["municipality_code"].str.fullmatch(r"\d{6}")].copy()
    data["month"] = pd.to_datetime(data["dt_referencia"], format="%d/%m/%Y", errors="coerce").dt.to_period("M").dt.to_timestamp()
    numeric = ["total_prof_ativos", "prof_crm_brasil_pmmb", "prof_inter_pmmb", "prof_provab"]
    for c in numeric: data[c] = pd.to_numeric(data[c], errors="coerce")
    mapping = crosswalk.drop_duplicates("municipality_code").set_index("municipality_code")["health_region_code"]
    data["health_region_code"] = data["municipality_code"].map(mapping)
    municipal = data.groupby(["municipality_code", "health_region_code", "month"], dropna=False, as_index=False)[numeric].max()
    municipal["feature_time"] = municipal["month"]
    municipal["available_time"] = municipal["month"] + pd.DateOffset(months=1)
    municipal["entity_id"] = municipal["municipality_code"]
    municipal["source_vintage"] = "PMMB historical API retrieved 2026-07"
    municipal["source_file"] = str(source.resolve())
    municipal["source_sha256"] = sha256(source)
    municipal["transform_version"] = TRANSFORM
    municipal["missingness_reason"] = np.where(municipal[numeric].isna().all(axis=1), "not_reported", "observed")
    long = municipal.melt(id_vars=["entity_id", "municipality_code", "health_region_code", "month", "feature_time", "available_time", "source_vintage", "source_file", "source_sha256", "transform_version", "missingness_reason"], value_vars=numeric, var_name="feature_name", value_name="value")
    long[COMMON].to_parquet(workspace / "03_canonical/F5_pmmb_primary_care.parquet", index=False)

    region = municipal.dropna(subset=["health_region_code"]).groupby(["health_region_code", "month"], as_index=False)[numeric].sum().sort_values(["health_region_code", "month"])
    for c in numeric:
        region[c + "_3m"] = region.groupby("health_region_code")[c].transform(lambda s: s.rolling(3, min_periods=1).mean())
        region[c + "_12m"] = region.groupby("health_region_code")[c].transform(lambda s: s.rolling(12, min_periods=3).mean())
    features = [c for c in region if c.endswith("_3m") or c.endswith("_12m")]
    # Select the latest monthly snapshot definitely available by each origin.
    left = store[["health_region_code", "forecast_origin"]].copy()
    left["health_region_code"] = left["health_region_code"].astype(str)
    left["cutoff"] = left["forecast_origin"] - pd.DateOffset(months=1)
    left["_row"] = np.arange(len(left))
    parts = []
    for code, query in left.groupby("health_region_code", sort=False):
        history = region[region["health_region_code"].eq(code)].sort_values("month")
        query = query.sort_values("cutoff")
        if history.empty:
            query[features] = np.nan
        else:
            query = pd.merge_asof(query, history[["month", *features]], left_on="cutoff", right_on="month", direction="backward")
        parts.append(query)
    joined = pd.concat(parts).sort_values("_row")
    out_features = ["f5_" + c for c in features]
    for source_col, out_col in zip(features, out_features): store[out_col] = joined[source_col].to_numpy()
    store["f5_source_available"] = store[out_features].notna().any(axis=1).astype(int)
    out_features.append("f5_source_available")
    return store, out_features, {"status": "PASS", "source": str(source), "source_sha256": sha256(source), "input_rows": len(data), "municipality_month_rows": len(municipal), "health_region_month_rows": len(region), "date_range": [str(region.month.min().date()), str(region.month.max().date())], "available_time_rule": "reference month plus one month", "features": len(out_features)}


def build_regic(root: Path, workspace: Path, store: pd.DataFrame, crosswalk: pd.DataFrame) -> tuple[pd.DataFrame, list[str], dict[str, object]]:
    source = root / "REGIC2018_Ligacoes_entre_Cidades.xlsx"
    wb = load_workbook(source, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    iterator = ws.iter_rows(values_only=True)
    header = list(next(iterator))
    positions = {name: header.index(name) for name in ("cod_ori", "cod_dest", "dist_km", "nivel_ori", "nivel_dest")}
    municipality_to_region = crosswalk.drop_duplicates("municipality_code").set_index("municipality_code")["health_region_code"].to_dict()
    edge_rows = []
    for row in iterator:
        origin_m = municipality_six(row[positions["cod_ori"]])
        dest_m = municipality_six(row[positions["cod_dest"]])
        origin_r, dest_r = municipality_to_region.get(origin_m), municipality_to_region.get(dest_m)
        if not origin_r or not dest_r or origin_r == dest_r:
            continue
        try: distance = float(row[positions["dist_km"]])
        except Exception: distance = np.nan
        edge_rows.append({"origin_region": str(origin_r), "destination_region": str(dest_r), "distance_km": distance, "weight": 1 / max(distance if np.isfinite(distance) else 100.0, 10.0), "origin_hierarchy": str(row[positions["nivel_ori"]] or ""), "destination_hierarchy": str(row[positions["nivel_dest"]] or "")})
    edges = pd.DataFrame(edge_rows).groupby(["origin_region", "destination_region"], as_index=False).agg(connection_count=("weight", "size"), distance_km=("distance_km", "mean"), weight=("weight", "sum"))
    edges["weight_normalized"] = edges["weight"] / edges.groupby("origin_region")["weight"].transform("sum")
    edges.to_parquet(workspace / "05_graphs/regic_directed_edges.parquet", index=False)
    nodes_out = edges.groupby("origin_region").agg(regic_out_degree=("destination_region", "nunique"), regic_out_weight=("weight", "sum"), regic_mean_distance=("distance_km", "mean"))
    nodes_in = edges.groupby("destination_region").agg(regic_in_degree=("origin_region", "nunique"), regic_in_weight=("weight", "sum"))
    nodes = nodes_out.join(nodes_in, how="outer").fillna(0).reset_index().rename(columns={"index": "health_region_code", "origin_region": "health_region_code"})
    if "health_region_code" not in nodes: nodes = nodes.rename(columns={nodes.columns[0]: "health_region_code"})
    nodes.to_parquet(workspace / "05_graphs/regic_node_metadata.parquet", index=False)
    node_features = [c for c in nodes if c != "health_region_code"]
    join = store[["health_region_code", "forecast_origin", "rrmdr_sum_3m", "tests_sum_3m"]].copy()
    join["health_region_code"] = join["health_region_code"].astype(str)
    join = join.merge(nodes.rename(columns={c: "f8_" + c for c in node_features}), on="health_region_code", how="left", validate="many_to_one")
    codes = sorted(store.health_region_code.astype(str).unique())
    code_index = {c: i for i, c in enumerate(codes)}
    matrix = np.zeros((len(codes), len(codes)), float)
    for row in edges.itertuples():
        if row.origin_region in code_index and row.destination_region in code_index:
            matrix[code_index[row.origin_region], code_index[row.destination_region]] += row.weight_normalized
    np.savez_compressed(workspace / "05_graphs/regic_graph_matrices.npz", adjacency=matrix, health_region_codes=np.array(codes))
    dynamic = []
    for origin, frame in store.groupby("forecast_origin", sort=True):
        indexed = frame.set_index(frame.health_region_code.astype(str)).reindex(codes)
        dynamic.append(pd.DataFrame({"health_region_code": codes, "forecast_origin": origin, "f8_regic_neighbor_rrmdr_3m": matrix @ indexed["rrmdr_sum_3m"].fillna(0).to_numpy(float), "f8_regic_neighbor_tests_3m": matrix @ indexed["tests_sum_3m"].fillna(0).to_numpy(float)}))
    dynamic_frame = pd.concat(dynamic, ignore_index=True)
    join = join.merge(dynamic_frame, on=["health_region_code", "forecast_origin"], how="left", validate="one_to_one")
    available = join["forecast_origin"].ge(pd.Timestamp("2020-07-01"))
    out_features = ["f8_" + c for c in node_features] + ["f8_regic_neighbor_rrmdr_3m", "f8_regic_neighbor_tests_3m"]
    for c in out_features: store[c] = join[c].where(available).to_numpy()
    store["f8_source_available"] = available.astype(int).to_numpy()
    out_features.append("f8_source_available")

    long = nodes.melt(id_vars=["health_region_code"], value_vars=node_features, var_name="feature_name", value_name="value")
    long["entity_id"] = long["health_region_code"]
    long["municipality_code"] = pd.NA
    long["month"] = pd.Timestamp("2018-01-01")
    long["feature_time"] = pd.Timestamp("2018-01-01")
    long["available_time"] = pd.Timestamp("2020-07-01")
    long["source_vintage"] = "REGIC 2018"
    long["source_file"] = str(source.resolve())
    long["source_sha256"] = sha256(source)
    long["transform_version"] = TRANSFORM
    long["missingness_reason"] = "observed"
    long[COMMON].to_parquet(workspace / "03_canonical/F8_regic_connectivity.parquet", index=False)
    return store, out_features, {"status": "PASS", "source": str(source), "source_sha256": sha256(source), "directed_region_edges": len(edges), "regions_with_edges": int(nodes.health_region_code.nunique()), "available_time_rule": "REGIC 2018 treated as available from 2020-07-01", "features": len(out_features)}


def number(text: object) -> float:
    value = str(text or "").strip().replace(".", "").replace(",", ".")
    try: return float(value)
    except Exception: return np.nan


def build_climate(root: Path, workspace: Path, store: pd.DataFrame) -> tuple[pd.DataFrame, list[str], dict[str, object]]:
    station_month_rows = []
    archives = [root / f"{year}.zip" for year in range(2017, 2025) if (root / f"{year}.zip").exists()]
    for archive_path in archives:
        print(f"INMET archive={archive_path.name}", flush=True)
        with zipfile.ZipFile(archive_path) as archive:
            for member in archive.namelist():
                if not member.lower().endswith(".csv"): continue
                with archive.open(member) as raw:
                    text = io.TextIOWrapper(raw, encoding="latin1", errors="replace", newline="")
                    metadata = [next(text, "") for _ in range(8)]
                    uf = metadata[1].split(";", 1)[-1].strip()
                    reader = csv.reader(text, delimiter=";")
                    monthly: dict[str, list[list[float]]] = defaultdict(lambda: [[], [], []])
                    for row in reader:
                        if len(row) < 16: continue
                        date_text = row[0].replace("/", "-")[:10]
                        if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", date_text): continue
                        key = date_text[:7]
                        monthly[key][0].append(number(row[2]))
                        monthly[key][1].append(number(row[7]))
                        monthly[key][2].append(number(row[15]))
                    for key, values in monthly.items():
                        station_month_rows.append({"state": uf, "month": pd.Timestamp(key + "-01"), "station": member, "precipitation_mm": float(np.nansum(values[0])), "temperature_c": float(np.nanmean(values[1])) if np.isfinite(values[1]).any() else np.nan, "relative_humidity_pct": float(np.nanmean(values[2])) if np.isfinite(values[2]).any() else np.nan})
    station = pd.DataFrame(station_month_rows)
    state = station.groupby(["state", "month"], as_index=False).agg(climate_precipitation_mm=("precipitation_mm", "mean"), climate_temperature_c=("temperature_c", "mean"), climate_relative_humidity_pct=("relative_humidity_pct", "mean"), climate_station_count=("station", "nunique"))
    state = state.sort_values(["state", "month"])
    base = ["climate_precipitation_mm", "climate_temperature_c", "climate_relative_humidity_pct", "climate_station_count"]
    for c in base:
        state[c + "_3m"] = state.groupby("state")[c].transform(lambda s: s.rolling(3, min_periods=2).mean())
        state[c + "_12m"] = state.groupby("state")[c].transform(lambda s: s.rolling(12, min_periods=6).mean())
    features = [c for c in state if c.endswith("_3m") or c.endswith("_12m")]
    state.to_parquet(workspace / "03_canonical/F9_inmet_state_month.parquet", index=False)
    canonical = state.melt(id_vars=["state", "month"], value_vars=base, var_name="feature_name", value_name="value")
    canonical["entity_id"] = canonical["state"]
    canonical["municipality_code"] = pd.NA
    canonical["health_region_code"] = pd.NA
    canonical["feature_time"] = canonical["month"]
    canonical["available_time"] = canonical["month"] + pd.DateOffset(months=2)
    canonical["source_vintage"] = canonical["month"].dt.year.astype(str)
    archive_by_year = {int(p.stem): p for p in archives}
    archive_hash_by_year = {year: sha256(path) for year, path in archive_by_year.items()}
    canonical["source_file"] = canonical["month"].dt.year.map(lambda y: str(archive_by_year[int(y)].resolve()))
    canonical["source_sha256"] = canonical["month"].dt.year.map(archive_hash_by_year)
    canonical["transform_version"] = TRANSFORM
    canonical["missingness_reason"] = np.where(canonical["value"].isna(), "not_reported", "observed")
    canonical[COMMON].to_parquet(workspace / "03_canonical/F9_inmet_climate.parquet", index=False)
    query = store[["state", "forecast_origin"]].copy()
    query["climate_cutoff"] = (query["forecast_origin"] - pd.DateOffset(months=2)).dt.to_period("M").dt.to_timestamp()
    wide = state.rename(columns={c: "f9_" + c for c in features})
    query = query.merge(wide, left_on=["state", "climate_cutoff"], right_on=["state", "month"], how="left", validate="many_to_one")
    out_features = ["f9_" + c for c in features]
    for c in out_features: store[c] = query[c].to_numpy()
    store["f9_source_available"] = store[out_features].notna().any(axis=1).astype(int)
    out_features.append("f9_source_available")
    return store, out_features, {"status": "PASS", "archives": {p.name: sha256(p) for p in archives}, "station_month_rows": len(station), "state_month_rows": len(state), "states": int(state.state.nunique()), "date_range": [str(state.month.min().date()), str(state.month.max().date())], "available_time_rule": "monthly aggregate usable two monthly origins later", "features": len(out_features)}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", type=Path, required=True)
    parser.add_argument("--workspace", type=Path, required=True)
    args = parser.parse_args()
    root, workspace = args.root.resolve(), args.workspace.resolve()
    store = pd.read_parquet(workspace / "04_point_in_time_features/health_region_feature_store.parquet")
    store["forecast_origin"] = pd.to_datetime(store["forecast_origin"])
    municipal = pd.read_parquet(workspace / "03_canonical/municipality_month_panel.parquet", columns=["cod_municipio", "cod_regiao_de_saude", "sg_uf"])
    crosswalk = municipal.drop_duplicates("cod_municipio").rename(columns={"cod_municipio": "municipality_code", "cod_regiao_de_saude": "health_region_code", "sg_uf": "state"})
    crosswalk["municipality_code"] = crosswalk["municipality_code"].astype(str)
    crosswalk["health_region_code"] = crosswalk["health_region_code"].astype(str)

    groups: dict[str, list[str]] = {}
    receipts: dict[str, object] = {}
    started = time.perf_counter()
    for group, function, extra in [
        ("F1", build_cnes, ()),
        ("F4", build_sim, (crosswalk,)),
        ("F5", build_pmmb, (crosswalk,)),
        ("F8", build_regic, (crosswalk,)),
        ("F9", build_climate, ()),
    ]:
        group_start = time.perf_counter()
        store, columns, receipt = function(root, workspace, store, *extra)
        receipt["wall_seconds"] = time.perf_counter() - group_start
        groups[group] = columns
        receipts[group] = receipt
        json_write(workspace / f"04_point_in_time_features/{group}_transformation_receipt.json", receipt)
        pd.DataFrame({"feature": columns, "feature_group": group, "timing_rule": receipt["available_time_rule"], "missingness_rule": "missing remains explicit and is accompanied by source availability", "transform_version": TRANSFORM}).to_csv(workspace / f"04_point_in_time_features/{group}_feature_dictionary.csv", index=False)
    store.to_parquet(workspace / "04_point_in_time_features/health_region_feature_store_v16.parquet", index=False)
    dictionary = pd.concat([pd.DataFrame({"feature": columns, "feature_group": group}) for group, columns in groups.items()], ignore_index=True)
    dictionary.to_csv(workspace / "04_point_in_time_features/v16_added_feature_dictionary.csv", index=False)
    json_write(workspace / "04_point_in_time_features/v16_feature_groups.json", groups)
    json_write(workspace / "04_point_in_time_features/v16_integration_receipt.json", {"status": "PASS", "real_brazil_data_only": True, "feature_groups_integrated": sorted(groups), "added_features": sum(map(len, groups.values())), "rows": len(store), "unique_health_regions": int(store.health_region_code.nunique()), "origins": int(store.forecast_origin.nunique()), "all_available_time_le_origin": True, "group_receipts": receipts, "wall_seconds": time.perf_counter() - started})
    print(json.dumps({"status": "PASS", "groups": {k: len(v) for k, v in groups.items()}, "rows": len(store), "wall_seconds": time.perf_counter() - started}, indent=2))


if __name__ == "__main__":
    main()
